# DB_Codebook_Generator_v2.py
# Code by Daniel Bañuelos — www.dandbc.mx/tools
# Try before using in professional workflows. Developed using Generative AI.

import json
import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from PIL import Image
import DaVinciResolveScript as dvr_script

SETTINGS_FILE = os.path.join(os.path.expanduser("~"), 'Documents', 'ResolveCodebook', 'codebook_settings.json')
os.makedirs(os.path.dirname(SETTINGS_FILE), exist_ok=True)

def frame_to_timecode(frame, fps):
    hours = int(frame / (3600 * fps))
    minutes = int((frame % (3600 * fps)) / (60 * fps))
    seconds = int((frame % (60 * fps)) / fps)
    frames = int(frame % fps)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}:{frames:02d}"

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_settings(settings):
    with open(SETTINGS_FILE, 'w') as f:
        json.dump(settings, f, indent=4)

def get_user_input(metadata_keys, default_settings):
    root = tk.Tk()
    root.title("Editorial Codebook Options")

    selected_fields = {f: tk.BooleanVar(value=default_settings.get('fields', {}).get(f, True)) for f in metadata_keys}
    frame_choice = tk.StringVar(value=default_settings.get("frame_choice", "middle"))
    thumb_size = tk.StringVar(value=default_settings.get("thumb_size", "small"))
    start_tc = tk.StringVar(value=default_settings.get("start_tc", "01:00:00:00"))
    custom_width = tk.StringVar(value=str(default_settings.get("custom_size", [320, 180])[0]))
    custom_height = tk.StringVar(value=str(default_settings.get("custom_size", [320, 180])[1]))
    delete_stills = tk.BooleanVar(value=default_settings.get("delete_stills", False))

    meta_frame = tk.LabelFrame(root, text="Metadata Fields")
    meta_frame.grid(row=0, column=0, padx=10, pady=5, sticky="w")
    field_frame = tk.Frame(meta_frame)
    field_frame.pack()

    def select_all():
        for var in selected_fields.values():
            var.set(True)

    def deselect_all():
        for var in selected_fields.values():
            var.set(False)

    btn_frame = tk.Frame(meta_frame)
    btn_frame.pack(anchor="w")
    tk.Button(btn_frame, text="Select All", command=select_all).pack(side="left")
    tk.Button(btn_frame, text="Deselect All", command=deselect_all).pack(side="left")

    for idx, f in enumerate(metadata_keys):
        tk.Checkbutton(field_frame, text=f, variable=selected_fields[f]).grid(row=idx % 20, column=idx // 20, sticky="w")

    thumb_frame = tk.LabelFrame(root, text="Thumbnail Options")
    thumb_frame.grid(row=1, column=0, padx=10, pady=5, sticky="w")
    tk.Label(thumb_frame, text="Size:").grid(row=0, column=0, sticky="w")
    for i, val in enumerate(["small", "big", "custom"]):
        tk.Radiobutton(thumb_frame, text=val.capitalize(), variable=thumb_size, value=val).grid(row=0, column=i+1, sticky="w")
    tk.Label(thumb_frame, text="Custom Width").grid(row=1, column=0, sticky="e")
    tk.Entry(thumb_frame, textvariable=custom_width, width=6).grid(row=1, column=1, sticky="w")
    tk.Label(thumb_frame, text="Custom Height").grid(row=1, column=2, sticky="e")
    tk.Entry(thumb_frame, textvariable=custom_height, width=6).grid(row=1, column=3, sticky="w")

    time_frame = tk.LabelFrame(root, text="Timeline Settings")
    time_frame.grid(row=2, column=0, padx=10, pady=5, sticky="w")
    tk.Label(time_frame, text="Frame to capture:").grid(row=0, column=0, sticky="w")
    for i, val in enumerate(["first", "middle", "last"]):
        tk.Radiobutton(time_frame, text=val.capitalize(), variable=frame_choice, value=val).grid(row=0, column=i+1, sticky="w")
    tk.Label(time_frame, text="Timeline Start TC:").grid(row=1, column=0, sticky="w")
    tk.Entry(time_frame, textvariable=start_tc).grid(row=1, column=1)

    misc_frame = tk.LabelFrame(root, text="Extras")
    misc_frame.grid(row=3, column=0, padx=10, pady=5, sticky="w")
    tk.Checkbutton(misc_frame, text="Delete all stills after export", variable=delete_stills).grid(row=0, column=0, sticky="w")

    def submit():
        root.quit()
        root.destroy()

    tk.Button(root, text="Generate Codebook", command=submit).grid(row=4, column=0, pady=10)
    root.mainloop()

    return {
        "fields": {f: var.get() for f, var in selected_fields.items()},
        "frame_choice": frame_choice.get(),
        "thumb_size": thumb_size.get(),
        "custom_size": (int(custom_width.get()), int(custom_height.get())),
        "start_tc": start_tc.get(),
        "delete_stills": delete_stills.get()
    }

def generate_codebook():
    resolve = dvr_script.scriptapp("Resolve")
    project = resolve.GetProjectManager().GetCurrentProject()
    timeline = project.GetCurrentTimeline()

    try:
        fps = float(project.GetSetting("timelineFrameRate"))
    except:
        print("❌ Could not retrieve timeline frame rate.")
        return

    items = []
    for track in range(1, timeline.GetTrackCount("video") + 1):
        items += timeline.GetItemListInTrack("video", track)

    if not items:
        messagebox.showwarning("No Clips", "No clips found in the timeline.")
        return

    sample_clip = items[0].GetMediaPoolItem().GetClipProperty()
    metadata_keys = sorted(sample_clip.keys())

    saved_settings = load_settings()
    opts = get_user_input(metadata_keys, saved_settings)
    timeline.SetStartTimecode(opts["start_tc"])
    save_settings(opts)

    root_dir = filedialog.askdirectory(title="Select Output Folder")
    if not root_dir:
        return

    project_name = project.GetName().replace(" ", "_")
    timeline_name = timeline.GetName().replace(" ", "_")
    default_filename = f"{project_name}_{timeline_name}_Editorial_Codebook"
    root = tk.Tk()
    root.withdraw()
    custom_filename = simpledialog.askstring("Output Filename", "Enter filename for your codebook:", initialvalue=default_filename)
    if not custom_filename:
        print("⚠️ User cancelled filename entry.")
        return

    file_base = custom_filename
    output_dir = os.path.join(root_dir, file_base)
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    fields = [k for k, v in opts["fields"].items() if v]
    headers = ["Thumbnail"] + fields
    ws.append(headers)

    thumbs_success = 0
    thumb_width = opts["custom_size"][0]
    thumb_height = opts["custom_size"][1]
    thumb_col_width = thumb_width / 7.5

    for i, clip in enumerate(items):
        mp_item = clip.GetMediaPoolItem()
        props = mp_item.GetClipProperty()
        row = []

        start = clip.GetStart()
        end = clip.GetEnd()
        duration = end - start
        frame = {"first": start, "middle": start + duration // 2, "last": end - 1}[opts["frame_choice"]]
        tc = frame_to_timecode(frame, fps)

        resolve.OpenPage("color")
        timeline.SetCurrentTimecode(tc)
        timeline.GrabStill()

        jpg_path = os.path.join(output_dir, f"thumb_{i}.jpg")
        success = project.ExportCurrentFrameAsStill(jpg_path)

        if success and os.path.exists(jpg_path):
            thumbs_success += 1
            if opts["thumb_size"] != "big":
                try:
                    img = Image.open(jpg_path)
                    img = img.resize((thumb_width, thumb_height))
                    img.save(jpg_path)
                except Exception as e:
                    print(f"⚠️ Failed to resize image: {e}")

            try:
                img = ExcelImage(jpg_path)
                img.anchor = f"A{i+2}"
                ws.add_image(img)
                ws.row_dimensions[i+2].height = thumb_height * 0.75
            except Exception as e:
                print(f"⚠️ Failed to embed image: {e}")
        else:
            print(f"⚠️ Failed to export thumbnail for clip {clip.GetName()}")

        for f in fields:
            row.append(props.get(f, ""))
        for col, val in enumerate(row, start=2):
            ws.cell(row=i+2, column=col, value=val)

    for col in range(2, len(headers)+2):
        col_letter = get_column_letter(col)
        max_len = max(len(str(ws.cell(row=row, column=col).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[col_letter].width = max(12, min(40, max_len + 2))
    ws.column_dimensions['A'].width = thumb_col_width

    xlsx_path = os.path.join(output_dir, f"{file_base}.xlsx")
    wb.save(xlsx_path)

    if opts.get("delete_stills"):
        try:
            resolve.OpenPage("color")
            gallery = project.GetGallery()
            album = gallery.GetCurrentStillAlbum()
            stills = album.GetStills()
            if stills:
                success = album.DeleteStills(stills)
                if success:
                    print("✅ All stills have been successfully deleted.")
                else:
                    print("⚠️ Failed to delete stills.")
            else:
                print("ℹ️ No stills found in album.")
        except Exception as e:
            print(f"⚠️ Could not delete stills: {e}")

    messagebox.showinfo(
        "Codebook Complete",
        f"✅ Editorial Codebook saved to:\n{xlsx_path}\n\n🎬 Clips processed: {len(items)}\n🖼️ Thumbnails exported: {thumbs_success}\n\nCode by Daniel Bañuelos\nwww.dandbc.mx/tools\nHappy Editing!"
    )

if __name__ == "__main__":
    generate_codebook()
