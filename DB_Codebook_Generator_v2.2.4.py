# DB_Codebook_Generator_v2.2_FINAL.py
# Developed by Daniel Bañuelos – www.dandbc.mx/tools
# License: BSD 3-Clause License
# Try before using in professional workflows. Developed using Generative AI.

import sys
import os
from sys import platform
if platform == "linux" or platform == "linux2":
    Resolve_Loc = '/opt/resolve/Developer/Scripting/Modules'
elif platform == "darwin":
    Resolve_Loc = '/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules'
elif platform == "win32":
    Resolve_Loc = r'C:\\ProgramData\\Blackmagic Design\\DaVinci Resolve\\Support\\Developer\\Scripting\\Modules'
else:
    print("Unsupported platform")
    Resolve_Loc = ""
sys.path.insert(1, Resolve_Loc)

import DaVinciResolveScript as dvr_script
import sys, os
from sys import platform
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import json

# Setup path for Resolve scripting
if platform == "darwin":
    Resolve_Loc = '/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules'
elif platform == "win32":
    Resolve_Loc = r'C:\\ProgramData\\Blackmagic Design\\DaVinci Resolve\\Support\\Developer\\Scripting\\Modules'
elif platform.startswith("linux"):
    Resolve_Loc = '/opt/resolve/Developer/Scripting/Modules'
else:
    print("❌ Unsupported platform."); sys.exit()
sys.path.insert(1, Resolve_Loc)

SETTINGS_FILE = os.path.join(os.path.expanduser("~"), 'Documents', 'ResolveCodebook', 'codebook_settings.json')
os.makedirs(os.path.dirname(SETTINGS_FILE), exist_ok=True)

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_settings(settings):
    with open(SETTINGS_FILE, 'w') as f:
        json.dump(settings, f, indent=4)

def frame_to_timecode(frame, fps):
    h = int(frame / (3600 * fps))
    m = int((frame % (3600 * fps)) / (60 * fps))
    s = int((frame % (60 * fps)) / fps)
    f = int(frame % fps)
    return f"{h:02d}:{m:02d}:{s:02d}:{f:02d}"

def get_user_input(metadata_keys, default_settings):
    root = tk.Tk()
    root.title("DB Codebook Generator v2.2")

    preferred_order = [
        "Clip Name", "Reel Name", "Start TC", "End TC", "Duration", "Good Take",
        "Scene", "Take", "Roll/Card", "Resolution", "Video Codec", "Audio Ch",
        "Audio Bit Depth", "Sample Rate", "Audio Offset", "Clip Color"
    ]

    ordered_keys = []
    added = set()
    for key in preferred_order:
        for real_key in metadata_keys:
            if key.lower().replace(" ", "") == real_key.lower().replace(" ", ""):
                ordered_keys.append(real_key)
                added.add(real_key)
                break
    ordered_keys += sorted(k for k in metadata_keys if k not in added)

    selected_fields = {f: tk.BooleanVar(value=default_settings.get('fields', {}).get(f, True)) for f in ordered_keys}
    frame_choice = tk.StringVar(value=default_settings.get("frame_choice", "middle"))
    thumb_size = tk.StringVar(value=default_settings.get("thumb_size", "small"))
    start_tc = tk.StringVar(value=default_settings.get("start_tc", "01:00:00:00"))
    custom_width = tk.StringVar(value=str(default_settings.get("custom_size", [320, 180])[0]))
    custom_height = tk.StringVar(value=str(default_settings.get("custom_size", [320, 180])[1]))
    delete_stills = tk.BooleanVar(value=default_settings.get("delete_stills", False))

    meta_frame = tk.LabelFrame(root, text="Metadata Fields")
    meta_frame.grid(row=0, column=0, padx=10, pady=5, sticky="w")
    field_frame = tk.Frame(meta_frame)
    field_frame.pack()

    def select_all(): [var.set(True) for var in selected_fields.values()]
    def deselect_all(): [var.set(False) for var in selected_fields.values()]
    tk.Button(meta_frame, text="Select All", command=select_all).pack(side="left")
    tk.Button(meta_frame, text="Deselect All", command=deselect_all).pack(side="left")

    for idx, f in enumerate(ordered_keys):
        tk.Checkbutton(field_frame, text=f, variable=selected_fields[f]).grid(row=idx % 20, column=idx // 20, sticky="w")

    thumb_frame = tk.LabelFrame(root, text="Thumbnail Options")
    thumb_frame.grid(row=1, column=0, padx=10, pady=5, sticky="w")
    for i, val in enumerate(["small", "big", "custom"]):
        tk.Radiobutton(thumb_frame, text=val.capitalize(), variable=thumb_size, value=val).grid(row=0, column=i, sticky="w")
    tk.Label(thumb_frame, text="Custom Width").grid(row=1, column=0)
    tk.Entry(thumb_frame, textvariable=custom_width, width=6).grid(row=1, column=1)
    tk.Label(thumb_frame, text="Height").grid(row=1, column=2)
    tk.Entry(thumb_frame, textvariable=custom_height, width=6).grid(row=1, column=3)

    time_frame = tk.LabelFrame(root, text="Timeline Settings")
    time_frame.grid(row=2, column=0, padx=10, pady=5, sticky="w")
    for i, val in enumerate(["first", "middle", "last"]):
        tk.Radiobutton(time_frame, text=val.capitalize(), variable=frame_choice, value=val).grid(row=0, column=i, sticky="w")
    tk.Label(time_frame, text="Start Timecode").grid(row=1, column=0)
    tk.Entry(time_frame, textvariable=start_tc).grid(row=1, column=1)

    misc_frame = tk.LabelFrame(root, text="Extras")
    misc_frame.grid(row=3, column=0, padx=10, pady=5, sticky="w")
    tk.Checkbutton(misc_frame, text="Delete stills after export", variable=delete_stills).grid(row=0, column=0)

    tk.Button(root, text="Generate Codebook", command=lambda: (root.quit(), root.destroy())).grid(row=4, column=0, pady=10)
    root.mainloop()

    return {
        "fields": {f: var.get() for f, var in selected_fields.items()},
        "frame_choice": frame_choice.get(),
        "thumb_size": thumb_size.get(),
        "custom_size": (int(custom_width.get()), int(custom_height.get())),
        "start_tc": start_tc.get(),
        "delete_stills": delete_stills.get()
    }

def generate_codebook():
    resolve = dvr_script.scriptapp("Resolve")
    project = resolve.GetProjectManager().GetCurrentProject()
    timeline = project.GetCurrentTimeline()
    fps = float(project.GetSetting("timelineFrameRate"))

    clips = []
    for t in range(1, timeline.GetTrackCount("video") + 1):
        clips.extend(timeline.GetItemListInTrack("video", t))

    first_valid = next((c for c in clips if c.GetMediaPoolItem()), None)
    props_example = first_valid.GetMediaPoolItem().GetClipProperty() if first_valid else {}
    metadata_keys = sorted(props_example.keys())

    opts = get_user_input(metadata_keys, load_settings())
    save_settings(opts)
    timeline.SetStartTimecode(opts["start_tc"])

    folder = filedialog.askdirectory()
    if not folder: return

    default_name = f"{project.GetName().replace(' ', '_')}_{timeline.GetName().replace(' ', '_')}_Codebook"
    root = tk.Tk(); root.withdraw()
    filename = simpledialog.askstring("Filename", "Enter filename:", initialvalue=default_name)
    if not filename: return

    out_dir = os.path.join(folder, filename)
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    fields = [f for f, v in opts["fields"].items() if v]
    ws.append(["Thumbnail"] + fields)

    processed = set()
    thumb_w, thumb_h = opts["custom_size"]

    for i, clip in enumerate(clips):
        mp_item = clip.GetMediaPoolItem()
        if not mp_item:
            continue
        try:
            props = mp_item.GetClipProperty()
        except:
            continue
        frame = {"first": clip.GetStart(), "middle": clip.GetStart() + (clip.GetEnd()-clip.GetStart())//2, "last": clip.GetEnd()-1}[opts["frame_choice"]]
        key = f"{mp_item.GetName()}_{frame}"
        if key in processed:
            continue
        processed.add(key)

        tc = frame_to_timecode(frame, fps)
        resolve.OpenPage("color")
        timeline.SetCurrentTimecode(tc)
        timeline.GrabStill()
        jpg_path = os.path.join(out_dir, f"thumb_{i}.jpg")
        if not project.ExportCurrentFrameAsStill(jpg_path): continue

        try:
            img = Image.open(jpg_path).resize((thumb_w, thumb_h))
            img.save(jpg_path)
            ex_img = ExcelImage(jpg_path)
            ex_img.anchor = f"A{i+2}"
            ws.add_image(ex_img)
            ws.row_dimensions[i+2].height = thumb_h * 0.75
        except: pass

        row = []
        for f in fields:
            if f.lower().replace(" ", "") == "clipcolor":
                color_name = getattr(clip, "GetClipColor", lambda: None)()
                row.append(color_name or "")
            else:
                row.append(props.get(f, ""))
        for c, v in enumerate(row, 2):
            ws.cell(row=i+2, column=c, value=v)

    for col in range(2, len(fields)+2):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 20
    ws.column_dimensions["A"].width = thumb_w / 7.5

    xlsx_path = os.path.join(out_dir, f"{filename}.xlsx")
    wb.save(xlsx_path)

    if opts["delete_stills"]:
        try:
            gallery = project.GetGallery()
            album = gallery.GetCurrentStillAlbum()
            stills = album.GetStills()
            if stills:
                album.DeleteStills(stills)
        except Exception as e:
            print(f"⚠️ Error deleting stills: {e}")

    messagebox.showinfo("Done", f"✅ Codebook saved to:\n{xlsx_path}\n\nDeveloped by Daniel Bañuelos\nwww.dandbc.mx/tools\nTry before using in professional workflows. Developed using Generative AI.")

if __name__ == "__main__":
    generate_codebook()
